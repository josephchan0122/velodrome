from configurations import importer  # noqa: E402 isort:skip
if not importer.installed:  # noqa: E402 isort:skip
    import configurations
    configurations.setup()
from collections import Counter
import datetime as dt
import itertools
import json
import logging
from random import randint
import string
import uuid

import celery
from concurrency.exceptions import RecordModifiedError
from django.apps import apps
from django.conf import settings
from django.contrib.gis.db.models import Extent
from django.core.exceptions import MultipleObjectsReturned, ValidationError
from django.core.mail import send_mail
from django.db import IntegrityError, models, transaction
from django.db.models import Value, signals
from django.db.models.aggregates import Max
from django.db.models.expressions import Case, DurationValue, F, When
from django.db.models.query import Q
from django.template.loader import get_template
from django.utils import timezone
from django.utils.crypto import get_random_string
from django_redis import get_redis_connection
from geopy.distance import geodesic
from humanize import naturaldelta
from openpyxl import load_workbook
from pinax.stripe.models import Plan
from reversion import revisions
import stripe


class Celery(celery.Celery):

    def on_configure(self):
        try:
            settings.RAVEN_DSN
        except AttributeError:
            pass
        else:
            from raven.contrib.celery import (
                register_logger_signal, register_signal,
            )
            from raven.contrib.django.models import get_client

            client = get_client()

            # register a custom filter to filter out duplicate logs
            register_logger_signal(client)

            # hook into the Celery error handler
            register_signal(client)


app = Celery('velodrome')
app.config_from_object('django.conf:settings', namespace='CELERY')
app.autodiscover_tasks()

from django.contrib.auth import get_user_model  # noqa: E402 isort:skip

logger = logging.getLogger(__name__)


@app.on_after_configure.connect
def setup_periodic_tasks(sender, **kwargs):
    sender.add_periodic_task(dt.timedelta(hours=1),
                             alert_tracking_not_received)
    sender.add_periodic_task(dt.timedelta(minutes=1),
                             expire_outdated_reservation)
    sender.add_periodic_task(dt.timedelta(hours=12), alert_idle_bicycles)
    sender.add_periodic_task(dt.timedelta(hours=12), notify_idle_bmmr_tasks)
    sender.add_periodic_task(dt.timedelta(days=1),
                             deactivate_overdue_maintenance_rules)
    sender.add_periodic_task(dt.timedelta(hours=1), start_future_bmmr)
    sender.add_periodic_task(dt.timedelta(days=1),
                             stop_lost_bicycle_reported_alerts)
    sender.add_periodic_task(dt.timedelta(minutes=1, seconds=15),
                             timeout_expired_rental_sessions)
    sender.add_periodic_task(dt.timedelta(hours=3, minutes=randint(1, 60)),
                             renew_refresh_tokens)
    sender.add_periodic_task(dt.timedelta(minutes=10),
                             stop_zone_alerts)


@app.task(bind=True)
@transaction.atomic
def fb_bulk_bicycle_import(self, file_id, user_pk, org_pk):
    from velodrome.lock8.models import (
        Bicycle,
        BicycleModel,
        BicycleType,
        Organization,
        User,
        private_storage,
    )

    with revisions.create_revision():
        try:
            owner = User.objects.get(pk=user_pk)
            revisions.set_user(owner)
            revisions.set_comment('fb_bulk_bicycle_import')
            organization = Organization.objects.get(pk=org_pk)
            city_bike = BicycleType.objects.get(reference='city_bike')
            open_file = private_storage.open(file_id)
            report = Counter()
            failures = {}
            wb = load_workbook(open_file, data_only=True, read_only=True)
            for model_name in wb.sheetnames:
                model, _ = BicycleModel.objects.get_or_create(
                    organization=organization,
                    name=model_name,
                    defaults={'owner': owner, 'type': city_bike},
                )

                sheet = wb[model_name]
                for row in sheet.iter_rows(min_row=3):
                    if not row:
                        break
                    description, name = row[0].value, row[1].value
                    if not (description and name):
                        break
                    if Bicycle.objects.filter(
                            description=description,
                            name=name,
                            model=model,
                            organization=organization).exists():
                        report['skipped'] += 1
                        continue
                    Bicycle.objects.create(
                        description=description,
                        name=name,
                        model=model,
                        owner=owner,
                        organization=organization,
                    )
                    report['created'] += 1

        except Exception as e:
            logger.exception('Uncaught exception during xlsx parsing: %s', e)
            transaction.set_rollback(True)  # doom the transaction
            send_mail('FB Bicycle bulk import {} failed'.format(file_id),
                      'Could not import spreadsheet because of this error:'
                      ' {}'.format(str(e)),
                      settings.DEFAULT_FROM_EMAIL,
                      [owner.email],
                      fail_silently=False,
                      )
            return
        finally:
            # always delete the file after parsing
            private_storage.delete(file_id)

    template = get_template('email/bulk_bicycle_import.txt')
    message = template.render(context={'report': report, 'failures': failures})
    send_mail('Bulk Bicycle import {}'.format(file_id),
              message,
              settings.DEFAULT_FROM_EMAIL,
              [owner.email],
              fail_silently=False,
              )


@app.task()
@transaction.atomic
def bulk_lock_creation(user_pk, lower_range, upper_range,
                       organization_pk, production_settings=''):
    from velodrome.lock8.models import Bicycle, Lock, Organization, User
    owner = User.objects.get(pk=user_pk)
    if organization_pk is not None:
        organization = Organization.objects.get(pk=organization_pk)
    else:
        organization = Organization.objects.get(level=0)
    total_counter = itertools.count()
    total_skipped = itertools.count()
    total_error = itertools.count()
    template = get_template('email/bulk_lock_creation.txt')
    try:
        for counter in range(lower_range, upper_range):
            loop_counter = itertools.count()
            while True:
                try:
                    with transaction.atomic():
                        lock = Lock(
                            organization=organization,
                            owner=owner,
                            counter=counter,
                            serial_number=production_settings + str(counter),
                            imei=get_random_string(
                                length=15,
                                allowed_chars=string.digits),
                            iccid=get_random_string(
                                length=20,
                                allowed_chars=string.digits),
                            sid=get_random_string(
                                length=32,
                                allowed_chars='abcdefABCDEF'),
                            bleid='TestLock{}{:_>10}'.format(
                                production_settings, counter),
                            randblock=get_random_string(length=2048),
                        )
                        lock.full_clean()
                except IntegrityError:
                    if next(loop_counter) == 10:
                        next(total_error)
                        break
                except ValidationError as exec_detail:
                    if ('Lock with this Serial number already exists.'
                            in exec_detail.message_dict.get('serial_number',
                                                            [])):
                        next(total_skipped)
                        break
                    if next(loop_counter) == 10:
                        next(total_error)
                        break
                else:
                    lock.provision()
                    Bicycle.objects.create(
                        name="Auto-bike #{}".format(counter),
                        organization=organization,
                        lock=lock,
                        owner=owner,
                    )
                    next(total_counter)
                    break
    except Exception as e:
        message = template.render(context={'exception': e})
    else:
        message = template.render(context={'total': next(total_counter),
                                           'skipped': next(total_skipped),
                                           'error': next(total_error)})
    send_mail('Bulk Lock creation report',
              message, settings.DEFAULT_FROM_EMAIL,
              [owner.email], fail_silently=False)


@app.task(bind=True)
@transaction.atomic
def bulk_lock_org_update(self, file_id, org_pk, user_pk):
    from velodrome.lock8.models import (
        Bicycle, Lock, LockStates, Organization, private_storage, User
    )

    report = Counter()
    error, failed = None, []
    template = get_template('email/bulk_lock_org_update.txt')
    owner = User.objects.get(pk=user_pk)
    new_organization = Organization.objects.get(pk=org_pk)
    open_file = private_storage.open(file_id)
    wb = load_workbook(open_file, data_only=True, read_only=True)

    try:
        only_sheet = wb.sheetnames[0]
        sheet = wb[only_sheet]
    except IndexError as err:
        error = 'CSV parsing failed: {}'.format(err)
    else:
        for row in sheet.iter_rows(min_row=2):
            serial_number = row[0].value
            EOF = (not serial_number)
            if EOF:
                break

            try:
                lock = Lock.objects.get(serial_number=serial_number)
            except Lock.DoesNotExist:
                failed.append('No Lock: {}'.format(serial_number))
                report['failed'] += 1
                continue
            if lock.state not in (LockStates.NEW.value,
                                  LockStates.PROVISIONED.value,
                                  LockStates.DECOMMISSIONED.value):
                failed.append('Invalid Lock {} with state: {}'.format(
                    serial_number, lock.state))
                report['failed'] += 1
                continue
            try:
                bicycle = lock.bicycle
            except Bicycle.DoesNotExist:
                pass
            else:
                failed.append('Device {} is assigned to bicycle {}'.format(
                    serial_number, bicycle.name))
                report['failed'] += 1
                continue
            if lock.organization == new_organization:
                report['skipped'] += 1
                continue
            lock.organization = new_organization
            lock.save()
            report['updated'] += 1

    private_storage.delete(file_id)
    message = template.render(context={
        'report': report,
        'failed': failed,
        'error': error,
        'destination': new_organization.name,
    })
    send_mail('Bulk lock organization report',
              message, settings.DEFAULT_FROM_EMAIL,
              [owner.email], fail_silently=False)


@app.task(bind=True)
def claim_axa_locks_from_spreadsheet_task(self, file_id, org_pk, user_pk):
    from velodrome.lock8.models import (
        AxaLock, Bicycle, Organization, private_storage, User,
    )

    report = Counter()
    error, failures = None, []
    template = get_template('email/claim_axa_locks.txt')
    owner = User.objects.get(pk=user_pk)
    organization = Organization.objects.get(pk=org_pk)
    open_file = private_storage.open(file_id)
    wb = load_workbook(open_file, data_only=True, read_only=True)

    try:
        only_sheet = wb.sheetnames[0]
        sheet = wb[only_sheet]
    except IndexError as err:
        error = 'CSV parsing failed: {}'.format(err)
    else:
        for row in sheet.iter_rows(min_row=2, max_col=3):
            qr_code = row[0].value.strip()
            _, lock_uid, claim_code = qr_code.split('-')
            if not claim_code:
                break
            try:
                serial_number = row[1].value
            except IndexError:
                serial_number = None

            with transaction.atomic():
                try:
                    try:
                        axa_lock = AxaLock.objects.get(uid=lock_uid)
                    except AxaLock.DoesNotExist:
                        axa_lock = AxaLock.objects.create(
                            organization=organization,
                            owner=owner,
                            uid=lock_uid,
                            claim_code_at_creation=uuid.UUID(hex=claim_code),
                        )
                        axa_lock.claim()
                        report['created'] += 1
                    if serial_number:
                        try:
                            with transaction.atomic():
                                bicycle = Bicycle.objects.filter(
                                    organization=organization,
                                    serial_number=serial_number,
                                    axa_lock__isnull=True).get()
                                bicycle.axa_lock = axa_lock
                                bicycle.full_clean()
                                bicycle.save()
                        except Exception as exc:
                            logger.exception(
                                'Exception while pairing axa locks')
                            report['failed'] += 1
                            failures.append(str(exc))
                        else:
                            report['paired'] += 1
                except Exception as exc:
                    logger.exception('Exception while claiming axa locks')
                    report['failed'] += 1
                    failures.append(str(exc))
                    transaction.set_rollback(True)
                    continue

    private_storage.delete(file_id)
    message = template.render(context={
        'report': report,
        'error': error,
        'failures': failures,
    })
    send_mail('Report about claimed Axa Locks from your spreadsheet',
              message, settings.DEFAULT_FROM_EMAIL,
              [owner.email], fail_silently=False)


@app.task(bind=True)
def declare_axa_lock_state_task(self, axa_lock_pk, state):
    from velodrome.lock8.models import (
        AxaLock,
    )

    axa_lock = AxaLock.objects.get(pk=axa_lock_pk)
    state_switchers = {
        'active': axa_lock.declare_active,
        'stored': axa_lock.declare_stored,
    }
    try:
        state_switchers[state]()
    except KeyError:
        logger.exception(
            f'Unable to find AxaLock method to switch it in {state} state'
        )


@app.task(bind=True)
@transaction.atomic
def generic_bulk_bicycle_import(self, file_id, user_pk, org_pk,
                                bicycle_model_pk):
    from velodrome.lock8.models import (
        Bicycle,
        BicycleModel,
        Organization,
        User,
        private_storage,
    )

    with revisions.create_revision():
        try:
            owner = User.objects.get(pk=user_pk)
            revisions.set_user(owner)
            revisions.set_comment('generic_bulk_bicycle_import')
            organization = Organization.objects.get(pk=org_pk)
            bicycle_model = BicycleModel.objects.get(pk=bicycle_model_pk)
            open_file = private_storage.open(file_id)
            report = Counter()
            failures = {}
            wb = load_workbook(open_file, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = [col.value for col in next(sheet.rows) if col.value]
                for row in sheet.iter_rows(min_row=2):
                    if not row:
                        break
                    change_dict = dict(zip(headers,
                                           (cell.value for cell in row)))
                    change_dict = {k: v for k, v in change_dict.items() if v}
                    if not change_dict:
                        break
                    if Bicycle.objects.filter(
                            model=bicycle_model,
                            organization=organization,
                            **change_dict).exists():
                        report['skipped'] += 1
                        continue
                    Bicycle.objects.create(
                        model=bicycle_model,
                        owner=owner,
                        organization=organization,
                        **change_dict
                    )
                    report['created'] += 1

        except Exception as e:
            logger.exception('Uncaught exception during xlsx parsing: %s', e)
            transaction.set_rollback(True)  # doom the transaction
            send_mail('Generic Bicycle bulk import {} failed'.format(file_id),
                      'Could not import spreadsheet because of this error:'
                      ' {}'.format(str(e)),
                      settings.DEFAULT_FROM_EMAIL,
                      [owner.email],
                      fail_silently=False,
                      )
            return
        finally:
            # always delete the file after parsing
            private_storage.delete(file_id)

    template = get_template('email/bulk_bicycle_import.txt')
    message = template.render(context={'report': report, 'failures': failures})
    send_mail('Bulk Bicycle import {}'.format(file_id),
              message,
              settings.DEFAULT_FROM_EMAIL,
              [owner.email],
              fail_silently=False,
              )


@app.task(bind=True)
@transaction.atomic
def assign_devices_to_bicycles(self, file_id, organization_pk, user_pk):
    from velodrome.lock8.models import (
        Bicycle, Lock, private_storage, User
    )

    report = Counter()
    error, failed = None, []
    template = get_template('email/assign_devices_to_bicycles.txt')
    owner = User.objects.get(pk=user_pk)
    open_file = private_storage.open(file_id)
    wb = load_workbook(open_file, data_only=True, read_only=True)
    try:
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
    except IndexError as err:
        error = 'CSV parsing failed: {}'.format(err)
    else:
        for row in itertools.islice(sheet.rows, 1, None, None):
            if not row:
                break

            bicycle_serial_number, lock_serial_number = (cell.value for cell
                                                         in row[0:2])
            if not (bicycle_serial_number or lock_serial_number):
                break
            try:
                bicycle = Bicycle.objects.get(
                    serial_number=bicycle_serial_number,
                    organization__pk=organization_pk,
                )
            except Bicycle.DoesNotExist:
                failed.append('No Bicycle: {}'.format(bicycle_serial_number))
                report['failed'] += 1
                continue
            except MultipleObjectsReturned:
                failed.append(
                    'Duplicate Bicycles with same serial_number: {}'.format(
                        bicycle_serial_number))
                report['failed'] += 1
                continue
            try:
                lock = Lock.objects.get(
                    serial_number=lock_serial_number,
                )
            except Lock.DoesNotExist:
                failed.append('No Lock: {}'.format(lock_serial_number))
                report['failed'] += 1
                continue
            if bicycle.lock == lock:
                failed.append('Lock {} already assigned to  {}'.format(
                    lock_serial_number, bicycle_serial_number))
                report['skipped'] += 1
                continue
            bicycle.lock = lock
            try:
                bicycle.full_clean()
                bicycle.save()
            except ValidationError as exc:
                failed.append('Inconsistent data for bicycle #{} and lock #{}:'
                              ' {}'.format(bicycle_serial_number,
                                           lock_serial_number, exc))
                report['failed'] += 1
            else:
                report['updated'] += 1

    private_storage.delete(file_id)
    message = template.render(context={
        'report': report,
        'failed': failed,
        'error': error,
    })
    send_mail('Assign Devices to Bicycles report.',
              message, settings.DEFAULT_FROM_EMAIL,
              [owner.email], fail_silently=False)


@transaction.atomic
def send_user_email(user_pk, action_name, *args):
    with revisions.create_revision():
        User = get_user_model()
        user = User.objects.get(pk=user_pk)
        revisions.set_user(user)
        revisions.set_comment('send_user_email: %s' % (action_name,))
        getattr(user, action_name)(*args)


@app.task()
def send_welcome_email_task(user_pk):
    send_user_email(user_pk, 'send_welcome_email')


@app.task()
def send_activation_email_task(user_pk):
    send_user_email(user_pk, 'send_activation_email')


@app.task()
def send_password_reset_email_task(user_pk):
    send_user_email(user_pk, 'send_password_reset_email')


@app.task()
def send_suspicious_registration_email_task(email, organization_pk):
    from velodrome.lock8.models import Organization
    from velodrome.lock8.utils import send_email
    organization = Organization.objects.get(pk=organization_pk)
    subject = (f'{organization.name} -'
               f' {settings.SUSPICIOUS_REGISTRATION_EMAIL_SUBJECT}')
    send_email(
        subject,
        [email],
        'email/user_suspicious_registration.txt',
        template_html='email/user_suspicious_registration.html',
        context={
            'organization_name': organization.name,
            'support_email': organization.get_preference('support_email')})


@app.task()
def send_invitation_email_task(invitation_pk, base_url):
    from velodrome.lock8.models import Invitation

    with revisions.create_revision():
        revisions.set_comment('send_invitation_email_task')
        invitation = Invitation.objects.get(pk=invitation_pk)
        invitation.send_invitation_email(base_url)


@app.task()
@transaction.atomic
def send_support_email_task(support_ticket_pk):
    from velodrome.lock8.models import SupportTicket

    with revisions.create_revision():
        revisions.set_comment('send_support_email_task')
        support_ticket = SupportTicket.objects.get(pk=support_ticket_pk)
        support_ticket.send_support_email()


@app.task()
@transaction.atomic
def sendable_task(pk, app_label, model_name):
    from django.apps import apps

    model = apps.get_model(app_label, model_name)
    with revisions.create_revision():
        sendable = model.objects.get(pk=pk)
        logger.info('Sending task: %s.%s: %r',
                    app_label, model_name, sendable)
        revisions.set_comment('sendable_task')
        sendable.send()


@app.task()
@transaction.atomic
def batched_alert_tracking_not_received(lock_ids):
    from velodrome.lock8.models import (
        Affiliation, Alert, Lock, maybe_create_and_send_alert)

    alert_type = Alert.NO_TRACKING_RECEIVED_SINCE
    for lock in Lock.objects.filter(pk__in=lock_ids):
        message = 'More than {} since the last periodic update.'.format(
            naturaldelta(settings.DEFAULT_INACTIVE_DURATION))
        maybe_create_and_send_alert(lock, alert_type, message,
                                    default_roles=[Affiliation.FLEET_OPERATOR])


BATCH_SIZE = 100


@app.task()
@transaction.atomic
def alert_tracking_not_received():
    from velodrome.lock8.models import (
        Alert,
        BicycleStates,
        GenericStates,
        Lock,
        LockStates,
    )

    alert_type = Alert.NO_TRACKING_RECEIVED_SINCE
    hours_24 = timezone.now() - settings.DEFAULT_INACTIVE_DURATION
    last_activity_timestamp = Case(
        When(bicycle__public_tracking__timestamp__isnull=False,
             then='bicycle__public_tracking__timestamp'),
        When(bicycle__public_tracking__timestamp__isnull=True,
             then='bicycle__transitions__timestamp'),
        default=Value(hours_24))
    already_alerted = Q(
        alerts__alert_type=alert_type,
        alerts__state=GenericStates.NEW.value,
    )
    locks = (
        Lock.objects
        .exclude(already_alerted)
        .annotate(last_activity_timestamp=last_activity_timestamp)
        .filter(last_activity_timestamp__lte=hours_24,
                state=LockStates.ACTIVE.value,
                bicycle__state__in=(BicycleStates.AVAILABLE.value,
                                    BicycleStates.RESERVED.value,
                                    BicycleStates.RENTED.value))
    )

    count = locks.count()
    for begin in range(0, count, BATCH_SIZE):
        lock_ids = list(locks.values_list('pk', flat=True)[begin:BATCH_SIZE])
        batched_alert_tracking_not_received.s(lock_ids).delay()


@app.task
@transaction.atomic
def renew_refresh_tokens(offset=0):
    from refreshtoken.models import RefreshToken

    hit = False
    expiracy_time = timezone.now() - settings.JWT_REFRESH_TOKEN_MAX_DURATION
    predicate = Q(created__lt=expiracy_time)
    for rt in RefreshToken.objects.filter(predicate)[offset:offset+BATCH_SIZE]:
        rt.revoke()
        hit = True
    if hit:
        renew_refresh_tokens.s(offset=offset + BATCH_SIZE).delay()


@app.task()
@transaction.atomic
def expire_outdated_reservation():
    from velodrome.lock8.utils import Ago
    from velodrome.lock8.models import (Bicycle, BicycleStates, User,
                                        ReservationStates)

    admin = User.objects.get(username='root_admin')

    # FIXME use recursion instead of hardcoding
    # grand parents relationship
    pred = {
        'organization__renting_scheme__max_reservation_duration__isnull':
        False}
    parent_pred = {
        'organization__parent__renting_scheme'
        '__max_reservation_duration__isnull': False}
    grand_parent_pred = {
        'organization__parent__parent__renting_scheme'
        '__max_reservation_duration__isnull': False}
    bicycles = (
        Bicycle.objects
        .annotate(
            max_reservation_duration=Case(
                When(renting_scheme__max_reservation_duration__isnull=False,
                     then=Max('renting_scheme__max_reservation_duration')),
                When(then=Max('organization__renting_scheme'
                              '__max_reservation_duration'),
                     **pred
                     ),
                When(then=Max('organization__parent__renting_scheme'
                              '__max_reservation_duration'),
                     **parent_pred
                     ),
                When(
                    then=Max('organization__parent__parent__renting_scheme'
                             '__max_reservation_duration'),
                    **grand_parent_pred
                ),
                default=DurationValue(
                    settings.DEFAULT_MAX_RESERVATION_DURATION),
            )
        )
        .filter(
            state=BicycleStates.RESERVED.value,
            reservation__created__lte=Ago('max_reservation_duration'),
            reservation__state=ReservationStates.NEW.value,
        ).distinct()
    )
    with revisions.create_revision():
        revisions.set_comment('expire_outdated_reservation')
        for bicycle in bicycles:
            try:
                bicycle.expire_reservation(by=admin)
            except RecordModifiedError:
                continue
            logger.info('Expired reservation for {}'.format(bicycle))


@app.task()
@transaction.atomic
def timeout_expired_rental_sessions():
    """Expire rental session if there is no pricing_scheme
    attached to the rental session. Duration is controlled
    by max_inactive_rental_session_duration preference, which defaults
    to settings.DEFAULT_MAX_INACTIVE_RENTAL_SESSION_DURATION (10mn).
    """
    from velodrome.lock8.utils import Ago
    from velodrome.lock8.models import (
        Affiliation, Alert, Bicycle, BicycleStates, User, RentalSessionStates,
        maybe_create_and_send_alert)

    admin = User.objects.get(username='root_admin')

    bicycles = (
        Bicycle.objects
        .annotate(rental_session_timestamp=Case(
            When(rental_session__state=RentalSessionStates.NEW.value,
                 then='rental_session__created'),
            output_field=models.DateTimeField()
        ))
        .annotate(
            max_inactive_rental_session_duration=Case(
                When(organization__preference__max_inactive_rental_session_duration__isnull=False,  # noqa
                     then=Max('organization__preference__max_inactive_rental_session_duration')),  # noqa
                When(organization__parent__preference__max_inactive_rental_session_duration__isnull=False,  # noqa
                     then=Max('organization__parent__preference__max_inactive_rental_session_duration')),  # noqa
                When(organization__parent__parent__preference__max_inactive_rental_session_duration__isnull=False,  # noqa
                     then=Max('organization__parent__parent__preference__max_inactive_rental_session_duration')),  # noqa
                default=DurationValue(
                    settings.DEFAULT_MAX_INACTIVE_RENTAL_SESSION_DURATION),
            ),
            has_an_active_pricing_scheme=Case(
                When(rental_session__state=RentalSessionStates.NEW.value,
                     rental_session__subscription_plan__pricing_scheme__isnull=False,  # noqa
                     then=True),
                When(rental_session__state=RentalSessionStates.NEW.value,
                     rental_session__pricing_scheme__isnull=False,
                     then=True),
                default=False,
                output_field=models.BooleanField()
            ),
            reference_timestamp=Case(
                When(public_tracking__gps_timestamp__gte=F(
                    'rental_session_timestamp'),
                     then='public_tracking__gps_timestamp'),
                default=F('rental_session_timestamp'))
        )
        .filter(
            has_an_active_pricing_scheme=False,
            state=BicycleStates.RENTED.value,
            reference_timestamp__lte=Ago(
                'max_inactive_rental_session_duration'),
        ).distinct()
    )
    with revisions.create_revision():
        revisions.set_comment('timeout_expired_rental_sessions')
        for bicycle in bicycles:
            user = bicycle.active_rental_session.user
            alert_context = {'renter': {'uuid': str(user.uuid),
                                        'display_name': user.display_name}}
            try:
                bicycle.expire_rental_session(by=admin)
            except RecordModifiedError:
                continue
            logger.info('Expired rental_session for %s (%s)',
                        bicycle, alert_context)
            maybe_create_and_send_alert(
                bicycle, Alert.BICYCLE_LEFT_UNLOCKED,
                owner=admin, default_roles=[Affiliation.FLEET_OPERATOR],
                context=alert_context)


@app.task()
@transaction.atomic
def alert_idle_bicycles(cursor=0):
    from velodrome.lock8.models import (
        Alert,
        AlertStates,
        Affiliation,
        BaseTracking,
        Bicycle,
        BicycleStates,
        ReadonlyTracking,
        maybe_create_and_send_alert,
    )
    alert_type = Alert.BICYCLE_IDLE_FOR_TOO_LONG
    already_alerted = Q(
        alerts__alert_type=alert_type,
        alerts__state__in=(AlertStates.NEW.value, AlertStates.ESCALATED.value,
                           AlertStates.SILENCED.value)
    )
    end = cursor + BATCH_SIZE
    bicycles = (
        Bicycle.objects
        .exclude(already_alerted)
        .filter(state=BicycleStates.AVAILABLE.value)
        .order_by('pk')
        .select_related('organization')
    )[cursor:end]

    LIMIT_LENGTH_IN_METER = 50
    for bicycle in bicycles:
        idle_bicycle_duration = bicycle.organization.idle_bicycle_duration
        if idle_bicycle_duration is None:
            continue
        area = (ReadonlyTracking.objects
                .filter(bicycle_uuid=bicycle.uuid,
                        organization_uuid=bicycle.organization.uuid,
                        tracking_type=BaseTracking.GPS_LOCATION_MESSAGE,
                        timestamp__gte=timezone.now() - idle_bicycle_duration)
                .aggregate(area=Extent('point')))['area']

        # if there is not enough ReadonlyTrackings for the period, we consider
        # the bicycle to be stationary.
        if area is None or len(area) < 4:
            pass
        else:
            # Switch/correct lat/lon.
            lon1, lat1 = area[:2]
            lon2, lat2 = area[2:]
            if geodesic((lat1, lon1), (lat2, lon2)).m >= LIMIT_LENGTH_IN_METER:
                continue

        message = 'This Bicycle is idle for more than {}'.format(
            naturaldelta(bicycle.organization.idle_bicycle_duration))
        maybe_create_and_send_alert(
            bicycle, alert_type, message,
            default_roles=[Affiliation.FLEET_OPERATOR]
        )
    if bicycles:
        alert_idle_bicycles.s(cursor=end).delay()


@app.task()
@transaction.atomic
def update_or_create_remote_plan(subscription_plan_pk):
    from velodrome.lock8.models import SubscriptionPlan, sync_plan
    from velodrome.lock8.utils import disable_signal

    subplan = SubscriptionPlan.objects.get(pk=subscription_plan_pk)
    plan = subplan.plan
    stripe_id = subplan.stripe_id
    org = subplan.organization
    stripe_account = org.stripe_account
    if plan is None:
        try:
            with transaction.atomic():
                plan = subplan.plan = Plan.objects.create(
                    stripe_id=stripe_id,
                    stripe_account=stripe_account,
                    **subplan.get_kwargs_for_plan(False))
        except IntegrityError:
            plan = Plan.objects.get(stripe_id=stripe_id,
                                    stripe_account=stripe_account)
        else:
            with disable_signal(signals.post_save, sync_plan, subplan):
                subplan.save()
    else:
        for k, v in subplan.get_kwargs_for_plan(True).items():
            setattr(plan, k, getattr(subplan, k))
        plan.save()

    stripe_account_id = stripe_account.stripe_id
    try:
        stripe_plan = stripe.Plan.retrieve(
            stripe_id, stripe_account=stripe_account_id)
    except stripe.error.InvalidRequestError:
        statement_descriptor = subplan.statement_descriptor
        if statement_descriptor == '':
            statement_descriptor = None
        try:
            stripe.Plan.create(
                id=stripe_id,
                amount=subplan.cents,
                currency=plan.currency,
                interval=subplan.interval,
                interval_count=subplan.interval_count,
                name=subplan.name,
                statement_descriptor=statement_descriptor,
                trial_period_days=subplan.trial_period_days,
                stripe_account=stripe_account_id,
            )
        except stripe.error.InvalidRequestError as exc:
            logger.exception('update_or_create_remote_plan: %r', exc)
            error_data = exc.json_body
            err = error_data.get('error', {})
            if (err.get('type') == 'invalid_request_error' and
                    err.get('message') == 'Plan already exists.'):
                stripe_plan = stripe.Plan.retrieve(
                    stripe_id, stripe_account=stripe_account_id)
            else:
                raise exc
        else:
            assert subplan.plan is not None
            return

    # Update remote Stripe plan.
    stripe_plan_updated = False
    for k in subplan.STRIPE_FIELDS_UPDATABLE:
        v = getattr(plan, k)
        # misalignment between pinax.stripe and Stripe
        if v == '':
            v = None
        if not hasattr(stripe_plan, k) or getattr(stripe_plan, k) != v:
            setattr(stripe_plan, k, v)
            stripe_plan_updated = True
    if stripe_plan_updated:
        stripe_plan.save()


@app.task()
def generate_payment(rental_session_pk):
    from velodrome.lock8.models import RentalSession

    rental_session = (RentalSession.objects
                      .filter(pk=rental_session_pk)
                      .select_related(
                          'bicycle',
                          'user',
                          'bicycle__organization').
                      get())
    rental_session.process_payment()


@app.task()
@transaction.atomic
def notify_idle_bmmr_tasks():
    from django.db.models import DateTimeField, ExpressionWrapper, F
    from velodrome.lock8.models import Task, TaskStates

    is_recurring_time = Q(maintenance_rule__fixed_date__isnull=True)
    is_fixed_time = Q(maintenance_rule__recurring_time__isnull=True)
    not_notified = Q(notification_messages__isnull=True)
    not_distance_based = Q(maintenance_rule__distance__isnull=True)
    fixed_time = F('maintenance_rule__fixed_date')
    computed_time = ExpressionWrapper(
        F('created') + F('maintenance_rule__recurring_time'),
        output_field=DateTimeField()
    )
    completed_or_due_or_cancelled = (
        Q(state__in=(TaskStates.COMPLETED.value,
                     TaskStates.CANCELLED.value)) |
        Q(is_due=True)
    )

    time_tasks = (
        Task.objects
        .exclude(completed_or_due_or_cancelled)
        .filter(not_distance_based, not_notified)
        .annotate(due_date=Case(
            When(is_recurring_time, then=computed_time),
            When(is_fixed_time, then=fixed_time)
        ))
        .filter(due_date__lt=timezone.now())
    )
    for task in time_tasks:
        task.is_due = True
        task.save()
        task.send_async()

    distance_based = Q(maintenance_rule__distance__isnull=False)
    dist_tasks = (
        Task.objects
        .exclude(completed_or_due_or_cancelled)
        .filter(distance_based, not_notified)
    )
    for task in dist_tasks:
        if task.get_remaining_distance() == 0:
            task.is_due = True
            task.save()
            task.send_async()


@app.task()
@transaction.atomic
def create_missing_tasks_async(pk):
    from velodrome.lock8.models import BicycleModelMaintenanceRule as BMMR

    rule = BMMR.objects.get(pk=pk)
    rule.create_missing_tasks()


@app.task()
@transaction.atomic
def start_future_bmmr():
    from velodrome.lock8.models import (BicycleModelMaintenanceRule as BMMR,
                                        BicycleModelMaintenanceRuleStates)

    for rule_pk in (BMMR.objects
                    .filter(
                        start_date__isnull=False,
                        start_date__lte=timezone.now(),
                        state=BicycleModelMaintenanceRuleStates.ACTIVE.value,
                        task__isnull=True,
                    )
                    .values_list('pk', flat=True)):
        create_missing_tasks_async.s(rule_pk).delay()


@app.task()
@transaction.atomic
def deactivate_overdue_maintenance_rules():
    from velodrome.lock8.models import (
        BicycleModelMaintenanceRule as BMMR,
        BicycleModelMaintenanceRuleStates as BMMRStates
    )

    is_overdue = Q(fixed_date__isnull=False, fixed_date__lt=timezone.now())
    overdue_bmmrs = (
        BMMR.objects
        .exclude(state=BMMRStates.DEACTIVATED.value)
        .filter(is_overdue)
    )

    with revisions.create_revision():
        revisions.set_comment('deactivate_overdue_maintenance_rules')
        for bmmr in overdue_bmmrs:
            bmmr.deactivate()


@app.task()
@transaction.atomic
def publish_pusher_update(app_label, model_name, instance_pk,
                          for_state_leaving):
    from velodrome.lock8.dispatchers import (
        build_publisher_serialization, build_publisher_topics)

    Model = apps.get_model(app_label=app_label, model_name=model_name)
    qs = Model.objects.filter(pk=instance_pk)
    if hasattr(Model.objects, 'prefetch_active'):
        qs = qs.prefetch_active(['rental_session', 'reservation'])

    instance = qs.get()
    topics = build_publisher_topics(
        instance, for_state_leaving=for_state_leaving)

    redis = get_redis_connection('publisher')
    payload_cache = {}
    count = 0
    for (topic, private) in topics:
        cache_key = (instance_pk, private)
        try:
            sender, update_payload = payload_cache[cache_key]
        except KeyError:
            sender, update_payload = build_publisher_serialization(
                instance, private, for_state_leaving=for_state_leaving)
            payload_cache[cache_key] = (sender, update_payload)
        payload = {'sender': sender,
                   'topic': topic,
                   'message': update_payload}
        message = json.dumps(payload)
        redis.publish(topic, message)
        count += 1
    logger.info('publish_pusher_update: %d topics for %s(pk=%d, %r)',
                count, model_name, instance_pk, for_state_leaving)


@app.task()
@transaction.atomic
def stop_lost_bicycle_reported_alerts():
    from velodrome.lock8.models import Alert, AlertStates

    two_weeks_ago = timezone.now() - dt.timedelta(weeks=2)
    for alert in Alert.objects.filter(created__lt=two_weeks_ago,
                                      state__in=(
                                          AlertStates.NEW.value,
                                          AlertStates.ESCALATED.value,
                                          AlertStates.SILENCED.value,
                                      )).select_for_update():
        alert.stop()


@app.task()
@transaction.atomic
def stop_alerts_for_rental(bicycle_pk, renter_pk):
    from velodrome.lock8.models import Alert, AlertStates, Bicycle, User

    bicycle = Bicycle.objects.get(pk=bicycle_pk)
    alerts = bicycle.alerts.filter(
        alert_type=Alert.BICYCLE_LEFT_UNLOCKED,
        state__in=(AlertStates.NEW.value,
                   AlertStates.SILENCED.value))
    if alerts:
        renter = User.objects.get(pk=renter_pk)
        logger.info("rent: stopping %d alerts for bicycle=%s",
                    len(alerts), bicycle.uuid)
        for alert in alerts:
            alert.stop(by=renter)


@app.task()
@transaction.atomic
def debug_celery_log():
    from velodrome.lock8.utils import debug_log_error

    logger.error('debug_celery_log: test logging error')
    logger.warning('debug_celery_log: test logging warning')
    debug_log_error('via debug_celery_log')


@app.task()
@transaction.atomic
def start_zone_alert_thresholds(zone_pk: int):
    from velodrome.lock8.models import (
        Affiliation, Alert, Zone, maybe_create_and_send_alert, Bicycle,
        BicycleStates)
    zone = Zone.objects.prefetch_related('alerts').get(pk=zone_pk)
    if zone.has_thresholds:
        zone_predicate = Q(public_tracking__point__intersects=zone.polygon)
        state_predicate = Q(state=BicycleStates.AVAILABLE.value)
        bicycles_count = Bicycle.objects.filter(
            zone_predicate, state_predicate).count()
        alert_type = message = None
        if bicycles_count >= zone.high_threshold:
            alert_type = Alert.ZONE_HIGH_THRESHOLD_TRIGGERED
            message = f'Zone {zone.name} has a high threshold alert'
        elif bicycles_count <= zone.low_threshold:
            alert_type = Alert.ZONE_LOW_THRESHOLD_TRIGGERED
            message = f'Zone {zone.name} has a low threshold alert'
        if alert_type:
            context = {
                'amount': bicycles_count,
                'zone_uuid': str(zone.uuid),
                'zone_name': zone.name,
                'zone_type': zone.type
            }
            maybe_create_and_send_alert(
                zone, alert_type, message, context=context,
                default_roles=[Affiliation.FLEET_OPERATOR])


@app.task()
@transaction.atomic
def stop_zone_alerts():
    """
    stop_zone_alerts is called every 10 minutes
    """
    from velodrome.lock8.models import Alert, AlertStates

    created = Q(created__lt=timezone.now() - dt.timedelta(minutes=60))
    alert_type = Q(alert_type__in=(Alert.ZONE_LOW_THRESHOLD_TRIGGERED,
                                   Alert.ZONE_HIGH_THRESHOLD_TRIGGERED))
    alert_state = Q(state__in=(AlertStates.NEW.value,
                               AlertStates.ESCALATED.value,
                               AlertStates.SILENCED.value))
    for alert in Alert.objects.filter(
            alert_state,
            alert_type,
            created).order_by().select_for_update():
        alert.stop()
